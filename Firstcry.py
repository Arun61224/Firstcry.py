import streamlit as st
import pandas as pd
import io
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# --- Core Logic Functions (Same as before) ---

def calculate_sale_price(product_cost, target_profit, gst_rate, royalty_percent, flat_rate, tds_rate, tcs_rate):
    """(Backward Calculation) Calculates the required Sale Price."""
    try:
        gst_r = gst_rate / 100.0
        royalty_r = royalty_percent / 100.0
        numerator = target_profit + product_cost
        
        # Denominator logic derived from:
        # Profit = SP * (1 - flat - royalty - (tds_rate + gst_r*tcs_rate) / (1+gst_r)) - Cost
        # SP = (Profit + Cost) / ( (1 - flat - royalty) - (tds_rate + gst_r*tcs_rate) / (1+gst_r) )
        
        denominator = (1 - flat_rate - royalty_r) - ((tds_rate + gst_r * tcs_rate) / (1 + gst_r))
        
        if denominator <= 0: 
            return None # Profit is impossible as deductions are >= 100%
        
        return numerator / denominator
    except Exception: 
        return None

def calculate_payout(sale_price, product_cost, gst_rate, royalty_percent, flat_rate, tds_rate, tcs_rate):
    """(Forward Calculation) Calculates profit from a given Sale Price."""
    try:
        if sale_price <= 0:
            return { "Net_Profit": -product_cost, "Final_Settled_Amount": -product_cost, "Flat_Deduction_Amount": 0, "Royalty_Fee_Amount": 0, "TDS_Amount": 0, "TCS_Amount": 0, "Taxable_Amount": 0 }
            
        gst_r = gst_rate / 100.0
        royalty_r = royalty_percent / 100.0
        
        taxable_amount = sale_price / (1 + gst_r)
        gst_value = sale_price - taxable_amount
        
        flat_deduction = sale_price * flat_rate
        royalty_fee = sale_price * royalty_r
        tds_amount = taxable_amount * tds_rate
        tcs_amount = gst_value * tcs_rate # TCS is on the tax component (GST)
        
        total_deductions = flat_deduction + royalty_fee + tds_amount + tcs_amount
        final_settled_amount = sale_price - total_deductions
        net_profit = final_settled_amount - product_cost
        
        return {
            "Net_Profit": net_profit,
            "Final_Settled_Amount": final_settled_amount,
            "Flat_Deduction_Amount": flat_deduction,
            "Royalty_Fee_Amount": royalty_fee,
            "TDS_Amount": tds_amount,
            "TCS_Amount": tcs_amount,
            "Taxable_Amount": taxable_amount
        }
    except Exception: 
        return None

# --- Excel Helper Function (Same as before) ---

def to_excel(df, cols_order, highlight_col_name=None):
    """Converts DataFrame to Excel in-memory, with optional highlighting."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # Ensure only columns that exist in the dataframe are used
        final_cols = [col for col in cols_order if col in df.columns]
        df_to_save = df[final_cols].fillna(0)
        df_to_save.to_excel(writer, index=False, sheet_name='Sheet1')
        
        worksheet = writer.sheets['Sheet1']
        
        if highlight_col_name and highlight_col_name in final_cols:
            highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Yellow
            
            # Find the column index in the *final* list of columns
            col_index = final_cols.index(highlight_col_name) + 1
            col_letter = get_column_letter(col_index)
            
            for cell in worksheet[col_letter][1:]: # Skip header
                cell.fill = highlight_fill
                
    processed_data = output.getvalue()
    return processed_data

# --- Streamlit UI ---

st.set_page_config(layout="wide", page_title="Profit Calculator")
st.title("💰 Profit Calculator App")

# --- Global Deduction Settings (Sidebar) ---
st.sidebar.header("Global Deduction Settings")
flat_rate = st.sidebar.number_input("Flat Deduction (e.g., 0.42)", value=0.42, min_value=0.0, max_value=1.0, step=0.01)
tds_rate = st.sidebar.number_input("TDS (on Taxable) (e.g., 0.001)", value=0.001, min_value=0.0, max_value=1.0, step=0.001, format="%.3f")
tcs_rate = st.sidebar.number_input("TCS (on Tax) (e.g., 0.10)", value=0.10, min_value=0.0, max_value=1.0, step=0.01)

# --- Main App Tabs ---
tab1, tab2, tab3, tab4 = st.tabs(["Bulk Payout Checker", "Bulk Price Calculator", "Single Payout Checker", "Single Price Calculator"])

# --- TAB 1: Bulk Payout Checker (Unchanged) ---
with tab1:
    st.header("Bulk Payout Checker (Forward)")
    st.write("Upload file with `Given_Sale_Price` and `Cost` to find the `Net_Profit` for all products.")

    # 1. Download Template
    with st.expander("Step 1: Download Payout Template"):
        payout_template_df = pd.DataFrame({
            "Product_SKU": ["SKU-001", "SKU-002"],
            "Given_Sale_Price": [1045.00, 1500.00],
            "Product_Cost": [500.00, 750.00],
            "GST_Rate_Percent": [5, 12],
            "Royalty_Percent": [10, 0]
        })
        st.download_button(
            label="Download Payout Template",
            data=to_excel(payout_template_df, payout_template_df.columns),
            file_name="payout_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # 2. Upload File
    st.subheader("Step 2: Upload and Process File")
    uploaded_payout_file = st.file_uploader("Upload your Payout Template", type=["xlsx"], key="payout_uploader")

    if uploaded_payout_file:
        try:
            df = pd.read_excel(uploaded_payout_file)
            st.dataframe(df.head(), use_container_width=True)

            required_cols = ["Given_Sale_Price", "Product_Cost", "GST_Rate_Percent", "Royalty_Percent"]
            if not all(col in df.columns for col in required_cols): # Check against required_cols
                st.error(f"Input file must have columns: {', '.join(required_cols)}")
            else:
                # 3. Process File
                if st.button("Process Payout File", type="primary"):
                    with st.spinner("Calculating..."):
                        results_list = []
                        for _, row in df.iterrows():
                            payout_data = calculate_payout(
                                row["Given_Sale_Price"], row["Product_Cost"], row["GST_Rate_Percent"],
                                row["Royalty_Percent"], flat_rate, tds_rate, tcs_rate)
                            results_list.append(payout_data)

                        results_df = pd.DataFrame(results_list)
                        df = df.join(results_df)
                        
                        cols_to_round = ["Final_Settled_Amount", "Net_Profit", "Taxable_Amount", "Flat_Deduction_Amount", "Royalty_Fee_Amount", "TDS_Amount", "TCS_Amount"]
                        df[cols_to_round] = df[cols_to_round].round(2)
                        
                        st.session_state.processed_payout_df = df
                        st.success(f"Processing Complete! {len(df)} products processed.")

        except Exception as e:
            st.error(f"Error processing file: {e}")

    # 4. Download Results
    if "processed_payout_df" in st.session_state:
        st.subheader("Step 3: Download Results")
        st.dataframe(st.session_state.processed_payout_df.head(), use_container_width=True)
        
        cols_order = [
            "Product_SKU", "Given_Sale_Price", "Product_Cost", "GST_Rate_Percent", "Royalty_Percent",
            "Final_Settled_Amount", "Net_Profit", "Taxable_Amount",
            "Flat_Deduction_Amount", "Royalty_Fee_Amount", "TDS_Amount", "TCS_Amount"
        ]
        
        excel_data = to_excel(st.session_state.processed_payout_df, cols_order, highlight_col_name="Net_Profit")
        st.download_button(
            label="Download Payout Results (Highlighted)",
            data=excel_data,
            file_name="payout_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


# --- TAB 2: Bulk Price Calculator (MODIFIED) ---
with tab2:
    st.header("Bulk Price Calculator (Reverse)")
    st.write("Upload file with `Cost` and `Target_Net_Profit` to find the `Required_Sale_Price`.")

    # 1. Download Template
    with st.expander("Step 1: Download Price Template"):
        price_template_df = pd.DataFrame({
            "Product_SKU": ["SKU-001", "SKU-002"],
            "Product_Cost": [500.00, 750.00],
            "Target_Net_Profit": [100.00, 150.00],
            "GST_Rate_Percent": [5, 12],
            "MRP": [1899.00, 2499.00],
            "Royalty_Percent": [10, 0]
        })
        st.download_button(
            label="Download Price Template",
            data=to_excel(price_template_df, price_template_df.columns),
            file_name="price_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # 2. Upload File
    st.subheader("Step 2: Upload and Process File")
    uploaded_price_file = st.file_uploader("Upload your Price Template", type=["xlsx"], key="price_uploader")

    if uploaded_price_file:
        try:
            df = pd.read_excel(uploaded_price_file)
            st.dataframe(df.head(), use_container_width=True)

            required_cols = ["Product_Cost", "Target_Net_Profit", "GST_Rate_Percent", "MRP", "Royalty_Percent"]
            if not all(col in df.columns for col in required_cols): # Check against required_cols
                st.error(f"Input file must have columns: {', '.join(required_cols)}")
            else:
                # 3. Process File
                if st.button("Process Price File", type="primary"):
                    with st.spinner("Calculating..."):
                        sale_prices = []
                        for _, row in df.iterrows():
                            sp = calculate_sale_price(
                                row["Product_Cost"], row["Target_Net_Profit"], row["GST_Rate_Percent"],
                                row["Royalty_Percent"], flat_rate, tds_rate, tcs_rate)
                            sale_prices.append(sp)
                        df["Required_Sale_Price"] = sale_prices
                        
                        # --- MODIFICATION: Calculate Net Payout and other details ---
                        taxable, flat, royalty, tds, tcs, discounts, statuses, net_payouts = [], [], [], [], [], [], [], []
                        
                        for _, row in df.iterrows():
                            sp, mrp = row["Required_Sale_Price"], row["MRP"]
                            cost, target_profit = row["Product_Cost"], row["Target_Net_Profit"]
                            gst_r, royalty_r = row["GST_Rate_Percent"] / 100.0, row["Royalty_Percent"] / 100.0
                            
                            if sp is None or pd.isna(sp):
                                # Append 0 for all numerical calculation columns
                                [arr.append(0) for arr in (taxable, flat, royalty, tds, tcs, net_payouts)]
                                discounts.append(None)
                                statuses.append("Profit Not Possible")
                                continue
                            
                            # Calculate deduction details
                            taxable_amount, gst_value = sp / (1 + gst_r), sp - (sp / (1 + gst_r))
                            taxable.append(taxable_amount)
                            flat.append(sp * flat_rate)
                            royalty.append(sp * royalty_r)
                            tds.append(taxable_amount * tds_rate)
                            tcs.append(gst_value * tcs_rate)
                            
                            # Calculate Net Payout (Final Settled Amount)
                            # This is simply the cost + target profit
                            net_payouts.append(cost + target_profit)
                            
                            # Check status
                            if sp > mrp: 
                                statuses.append("Error: Sale Price > MRP")
                                discounts.append(None)
                            else: 
                                statuses.append("OK")
                                discounts.append(((mrp - sp) / mrp) * 100)
                        
                        # Add new columns to DataFrame
                        df["Taxable_Amount"] = taxable
                        df["Flat_Deduction_Amount"] = flat
                        df["Royalty_Fee_Amount"] = royalty
                        df["TDS_Amount"] = tds
                        df["TCS_Amount"] = tcs
                        df["Discount_Percent"] = discounts
                        df["Net_Payout_Amount"] = net_payouts # <-- ADDED
                        df["Status"] = statuses
                        
                        # Add Net_Payout_Amount to rounding list
                        cols_to_round = [
                            "Required_Sale_Price", "Taxable_Amount", "Flat_Deduction_Amount", 
                            "Royalty_Fee_Amount", "TDS_Amount", "TCS_Amount", 
                            "Discount_Percent", "Net_Payout_Amount"
                        ]
                        df[cols_to_round] = df[cols_to_round].round(2)
                        
                        st.session_state.processed_price_df = df
                        st.success(f"Processing Complete! {len(df)} products processed.")

        except Exception as e:
            st.error(f"Error processing file: {e}")

    # 4. Download Results
    if "processed_price_df" in st.session_state:
        st.subheader("Step 3: Download Results")
        st.dataframe(st.session_state.processed_price_df.head(), use_container_width=True)
        
        # --- MODIFIED: Updated column order for export ---
        cols_order = [
            "Product_SKU", "Product_Cost", "Target_Net_Profit", "GST_Rate_Percent", "MRP", "Royalty_Percent",
            "Required_Sale_Price", 
            "Net_Payout_Amount", # <-- MOVED HERE (as requested)
            "Discount_Percent", 
            "Taxable_Amount",
            "Flat_Deduction_Amount", "Royalty_Fee_Amount", "TDS_Amount", "TCS_Amount",
            "Status" # <-- Moved to the end
        ]
        
        excel_data = to_excel(st.session_state.processed_price_df, cols_order, highlight_col_name="Required_Sale_Price")
        st.download_button(
            label="Download Price Results (Highlighted)",
            data=excel_data,
            file_name="price_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# --- TAB 3: Single Payout Checker (Unchanged) ---
with tab3:
    st.header("Single Payout Checker (Forward)")
    st.write("Enter a `Sale Price` to see your `Net Profit` and all deductions.")
    
    with st.form("single_payout_form"):
        col1, col2 = st.columns(2)
        with col1:
            sp_entry = st.number_input("Given Sale Price (₹)", min_value=0.0, step=1.0)
            pc_entry = st.number_input("Product Cost (₹)", min_value=0.0, step=1.0)
        with col2:
            gst_entry = st.number_input("GST Rate (%)", min_value=0.0, step=1.0)
            roy_entry = st.number_input("Royalty (%)", min_value=0.0, value=0.0, step=1.0)
        
        submitted = st.form_submit_button("Calculate Payout", type="primary")

    if submitted:
        results = calculate_payout(sp_entry, pc_entry, gst_entry, roy_entry, flat_rate, tds_rate, tcs_rate)
        
        if results:
            st.subheader("Results")
            
            profit_color = "normal"
            if results["Net_Profit"] < 0: profit_color = "inverse"
            elif results["Net_Profit"] > 0: profit_color = "normal"

            col1, col2 = st.columns(2)
            col1.metric("Final Settled Amount", f"₹ {results['Final_Settled_Amount']:,.2f}")
            col2.metric("Net Profit", f"₹ {results['Net_Profit']:,.2f}", delta_color=profit_color)
            
            with st.expander("Show Deduction Details"):
                st.write(f"**Taxable Amount:** ₹ {results['Taxable_Amount']:,.2f}")
                st.write(f"**Flat Fee:** ₹ {results['Flat_Deduction_Amount']:,.2f}")
                st.write(f"**Royalty Fee:** ₹ {results['Royalty_Fee_Amount']:,.2f}")
                st.write(f"**TDS:** ₹ {results['TDS_Amount']:,.2f}")
                st.write(f"**TCS:** ₹ {results['TCS_Amount']:,.2f}")
        else:
            st.error("Calculation Error.")

# --- TAB 4: Single Price Calculator (Unchanged) ---
with tab4:
    st.header("Single Price Calculator (Reverse)")
    st.write("Enter your `Cost` and `Target Profit` to find the `Required Sale Price`.")
    
    with st.form("single_price_form"):
        col1, col2 = st.columns(2)
        with col1:
            pc_price_entry = st.number_input("Product Cost (₹)", min_value=0.0, step=1.0, key="sp_pc")
            profit_price_entry = st.number_input("Target Net Profit (₹)", min_value=0.0, step=1.0, key="sp_profit")
        with col2:
            gst_price_entry = st.number_input("GST Rate (%)", min_value=0.0, step=1.0, key="sp_gst")
            roy_price_entry = st.number_input("Royalty (%)", min_value=0.0, value=0.0, step=1.0, key="sp_roy")
        
        submitted_price = st.form_submit_button("Calculate Price", type="primary")

    if submitted_price:
        # Calculate the required sale price
        req_sp = calculate_sale_price(
            pc_price_entry, 
            profit_price_entry, 
            gst_price_entry, 
            roy_price_entry, 
            flat_rate, 
            tds_rate, 
            tcs_rate
        )
        
        if req_sp is not None:
            st.subheader("Results")
            
            # Primary Metric: Required Sale Price
            st.metric("Required Sale Price", f"₹ {req_sp:,.2f}", 
                        help="This is the Sale Price (MRP inclusive of all taxes) you need to set.")
            
            # --- Verification Step ---
            # Re-calculate the payout using the new Sale Price to show details
            results = calculate_payout(
                req_sp, 
                pc_price_entry, 
                gst_price_entry, 
                roy_price_entry, 
                flat_rate, 
                tds_rate, 
                tcs_rate
            )
            
            if results:
                st.write("---")
                st.markdown(f"**Verification for this Sale Price:**")
                
                col1, col2 = st.columns(2)
                col1.metric("Final Settled Amount", f"₹ {results['Final_Settled_Amount']:,.2f}")
                # The net profit should be very close (or identical) to the target profit
                col2.metric("Verified Net Profit", f"₹ {results['Net_Profit']:,.2f}", 
                            delta=f"₹ {results['Net_Profit'] - profit_price_entry:,.2f} vs Target")
                
                with st.expander("Show Deduction Details for this Price"):
                    st.write(f"**Taxable Amount:** ₹ {results['Taxable_Amount']:,.2f}")
                    st.write(f"**Flat Fee:** ₹ {results['Flat_Deduction_Amount']:,.2f}")
                    st.write(f"**Royalty Fee:** ₹ {results['Royalty_Fee_Amount']:,.2f}")
                    st.write(f"**TDS:** ₹ {results['TDS_Amount']:,.2f}")
                    st.write(f"**TCS:** ₹ {results['TCS_Amount']:,.2f}")
            
        else:
            st.error("Calculation Error: Profit Not Possible.")
            st.write("This usually means the deductions (Flat rate, Royalty, etc.) are too high, making the target profit impossible to achieve.")
